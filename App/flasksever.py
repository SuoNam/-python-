import time

from flask import Flask, request,render_template,make_response,session,redirect,send_file,jsonify
import sql
import os
import json
from openpyxl import load_workbook
import secrets
from itsdangerous import URLSafeSerializer
import sys
from datetime import datetime
import threading
import  schedule
app = Flask(__name__,template_folder='templates',static_folder='static')

secret_key = secrets.token_hex(16)#密钥
app.secret_key = secret_key
serializer = URLSafeSerializer(app.secret_key)

m_sql=sql.mysql()
m_sql.build_sql()#数据库建立
m_sql.build_student_test()#被试表
m_sql.build_table_pre_id()#预用户密码表
m_sql.build_table_id()#用户密码表
m_sql.build_table_experiment()#实验表
#m_sql.inser_id_password_admin((1,'356a192b7913b04c54574d18c28d46e6395428ab',))
port=sys.argv[1]
def test_job():
    m_sql.build_sql()
    m_sql.select_pre_id_login()
    print("test_begin!")
def heart_check():
    while True:
        test_job()
        time.sleep(60)
# def db_job():
#     m_sql = sql.mysql()
#     m_sql.build_sql()
#     d=m_sql.select_pre_id_login()
#     print(d)
#     print("定时任务开始")
#
# def scheduled_job():
#     schedule.every(6).hours.do(db_job)
#
# schedule.every().hour.do(scheduled_job)
#到达指定时间删除文件
def delete_file_at_time(file_path, target_time):
    # 获取当前时间
    current_time = datetime.now().date()
    # 如果当前时间超过了目标时间，则删除指定文件
    if current_time >= target_time:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"文件 {file_path} 已被删除！")
            return 1
        else:
            print(f"文件 {file_path} 不存在！")
            return 1
    else:
        return 0
def d_file_at_time(file_path,target_time):
    v=0
    target_time = datetime.strptime(target_time, '%Y-%m-%d').date()
    while v==0:
        v=delete_file_at_time(file_path,target_time)

@app.route('/')
def handle_root():#根目录访问登录界面
    resp = redirect('/login?_method=GET')
    return resp
@app.route('/releasedInformation',methods=['POST','GET'])#招募界面
def handle_recruit():
    return render_template('releasedInformation.html')
@app.route('/img/<path:file_name>', methods=['GET','POST'])#处理照片
def handle_img(file_name):
        return send_file(f'img/{file_name}', mimetype='image/jpeg')
@app.route('/register',methods=['GET','POST'])
def handle_register():#注册
    m_sql.build_sql()#数据库建立
    if request.method== 'GET':
        return render_template("register.html")
    if request.method=='POST':
        #获取数据
        name=request.form.get('name')#姓名
        id=request.form.get('id')#学号
        password_sha1=request.form.get('password')#密码
        phone_number=request.form.get('phone')#手机号
        group=request.form.get("group")#课题组
        mentor=request.form.get("mentor")#导师

        if m_sql.select_id_login(id,password_sha1)!=-1:
            resp = redirect('/login?_method=GET&message_type=success&message=账号已存在')
            return resp
        if m_sql.select_id_pre_id_lohin(id)!=[]:
            resp = redirect('/login?_method=GET&message_type=success&message=您的注册已经提交请耐心等待')
            return resp
        #添加进数据库
        m_sql.inser_pre_id_password(name,id,password_sha1,phone_number,group,mentor)
        resp = redirect('/login?_method=GET&message_type=success&message=请等待审批')
        return resp
@app.route('/query_audit',methods=['POST'])
def handle_query_audit():#获取审核
    if request.method == 'OPTIONS':
        data = {'message': 'OK!'}
        response = make_response(json.dumps(data))
        response.headers['Server'] = 'test/1.0'
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Cache-Control'] = 'no-store'
        response.headers['Access-Control-Allow-Headers'] = '*'
        response.headers['Access-Control-Allow-Methods'] = '*'
        return response
    if request.method == 'POST':
        # 获取预用户密码表
        # 接受返回的json
        packet = request.get_json()
        # 解析json
        page = packet['page']
        query_data = packet["filter"]
        # 获取待查询数据
        j_data = m_sql.select_pre_id_login() ##占行
        if query_data == {}:
            j_data = m_sql.select_pre_id_login() # 得到数据
        else:
            k_l = list(query_data.keys())
            for i in k_l:
                if i == "date":
                    print("error message")#占行
                    exit(0)
        # 数据分块
        length = len(j_data) // 10
        remainder= len(j_data) % 10
        if length != 0:
            if remainder != 0:
                data_l = [j_data[i:i + 10] for i in range(0, length * 10, 10)]
                remainder_d = j_data[length * 10:]
                if page != length + 1:
                    message = {"length": length + 1, "data": data_l[page - 1]}
                else:
                    message = {"length": length + 1, "data": remainder_d}
            else:
                data_l = [j_data[i:i + 10] for i in range(0, length * 10, 10)]
                message = {"length": length, "data": data_l[page - 1]}
        else:
            if remainder != 0:
                message={"length":1, "data":j_data}
            else:
                message = {"length": 0, "data":[]}


        response = make_response(json.dumps(message))
        response.headers['Server'] = 'test/1.0'
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Cache-Control'] = 'no-store'
        response.headers['Access-Control-Allow-Headers'] = '*'
        response.headers['Access-Control-Allow-Methods'] = '*'
        return response
@app.route('/audit',methods=['POST'])
def handle_audit():#审核
    data=request.get_json()
    id=data['id']
    m_data=m_sql.select_id_pre_id_lohin(id)
    try:
        m_sql.delete_pre_id_password(id)
    except:
        message = {'message': "删除数据出错","status":0, 'messageType': "error"}
        return jsonify(message)
    else:
        if data['pass']=='true':
            try:
                m_sql.inser_id_password(m_data)
            except:
                message = {'message': "增添数据出错", "status": 0, 'messageType': "error"}
                return jsonify(message)
            else:
                message = {'message': "通过成功", "status": 1, 'messageType': "success"}
                return jsonify(message)
        else:
            message = {'message':"拒绝成功", "status": 1, 'messageType': "success"}
            return jsonify(message)
@app.route('/login', methods=['POST','GET'])#登录窗口，开放网页
def handle_login():
 print("数据库准备建立")
 m_sql.build_sql()#数据库建立
 if request.method == 'GET':
     identify = session.get('id')
     if identify:
         return redirect('/interface/experimenterCenter.html?_method=GET&message_type=success&message=欢迎回来')
     else:
         return render_template('login.html')

 if request.method=='POST':
    id = request.form['id']
    signal=m_sql.select_id_login_signal(id)#在数据库中查找id获取身份
    if id=='':
        resp = redirect('/login?_method=GET&message_type=error&message=账户不存在')
        return resp
    password_sha256=request.form['password']
    result=m_sql.select_id_login(id,password_sha256)
    if result==1:
        print("LOGIN SUCCESS!")
        session['id'] = id
        signed_value = serializer.dumps(id)
        resp = redirect('/interface/experimenterCenter.html?_method=GET&message_type=success&message=登录成功')
        resp.set_cookie('id', id)
        resp.set_cookie('signal',signal[0])
        resp.set_cookie('sign_name', value=signed_value, secure=False)
        return resp
    elif result==0:
        resp= redirect('/login?_method=GET&message_type=error&message=密码错误')
        return resp
    else:
        resp = redirect('/login?_method=GET&message_type=error&message=账户不存在')
        return resp

@app.route('/logout')#登出
def logout():
    # 删除 session 中存储的用户名
    session.pop('id', None)
    # 创建响应对象，删除 cookie
    resp = redirect('/login?_method=GET&message_type=success&message=退出成功')
    resp.set_cookie('id', '', expires=0)
    return resp


@app.route('/interface/<path:file_name>', methods=['GET'])  # 挂网页
def handle_index(file_name):
        cookie_value = request.cookies.get('sign_name')#签证
        if cookie_value:
            try:
                original_value = serializer.loads(cookie_value)
            except Exception:
                resp = redirect('/logout?_method=GET&message_type=error&message=cookie异常1')
                return resp
        else:
            resp = redirect('/logout?_method=GET&message_type=error&message=cookie异常2')
            return resp
        flage = session.get('id')
        if flage==request.cookies.get('id') and flage!=None:
            if request.cookies.get('signal') == m_sql.select_id_login_signal(flage)[0]:
                    return render_template('{}'.format(file_name), username=flage)
            else:
                resp = redirect('/logout?_method=GET&message_type=warning&message=cookie异常3')
                return resp
        else:
            resp = redirect('/logout?_method=GET&message_type=warning&message=登录过期')
            return resp

@app.route('/query_records', methods=['OPTIONS', 'POST'])  # 处理options请求以及  实验记录查询post
def handle_sumit():
        if request.method == 'OPTIONS':
            data = {'message': 'OK!'}
            response = make_response(json.dumps(data))
            response.headers['Server'] = 'test/1.0'
            response.headers['Content-Type'] = 'application/json; charset=utf-8'
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Cache-Control'] = 'no-store'
            response.headers['Access-Control-Allow-Headers'] = '*'
            response.headers['Access-Control-Allow-Methods'] = '*'
            return response
        if request.method == 'POST':
            # 接受返回的json
            packet = request.get_json()
            #解析json
            page=packet['page']
            query_data=packet["filter"]
            #获取待查询数据
            j_data = m_sql.select_sql_id_test(session['id'])##占行
            if query_data=={}:
                j_data = m_sql.select_sql_id_test(session['id'])#得到数据
            else:
                k_l=list(query_data.keys())
                for i in k_l:
                    if i=="date":
                        j_data =m_sql.select_sql_id_test_time(session['id'],query_data['date']['startDate'],query_data['date']['endDate'])
            #数据分块
            length = len(j_data) // 10
            remainder = len(j_data) % 10
            if length != 0:
                if remainder != 0:
                    data_l = [j_data[i:i + 10] for i in range(0, length * 10, 10)]
                    remainder_d = j_data[length * 10:]
                    if page != length + 1:
                        message = {"length": length + 1, "data": data_l[page - 1]}
                    else:
                        message = {"length": length + 1, "data": remainder_d}
                else:
                    data_l = [j_data[i:i + 10] for i in range(0, length * 10, 10)]
                    message = {"length": length, "data": data_l[page - 1]}
            else:
                if remainder != 0:
                    message = {"length": 1, "data": j_data}
                else:
                    message = {"length": 0, "data": []}

            response = make_response(json.dumps(message))
            response.headers['Server'] = 'test/1.0'
            response.headers['Content-Type'] = 'application/json; charset=utf-8'
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Cache-Control'] = 'no-store'
            response.headers['Access-Control-Allow-Headers'] = '*'
            response.headers['Access-Control-Allow-Methods'] = '*'
            return response

@app.route('/upload', methods=['POST'])  # 处理文件传输
def handle_upload():
        if 'file' not in request.files:
            resp = redirect('/interface/upload.html?_method=GET&message_type=error&message=文件传输错误')
            return resp
        file = request.files['file']
        file_name = session.get('id')
        if file.filename == '':
            resp = redirect('/interface/upload.html?_method=GET&message_type=error&message=文件传输错误')
            return resp
        UPLOAD_FOLDER = 'file'
        app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
        file.save(os.path.join(app.config['UPLOAD_FOLDER'],f"{file_name}.xlsx"))
        #设置处理xlxs，
        try:
            wb = load_workbook(filename="./file/" + f"{file_name}.xlsx")
            ws = wb.active
        except:
            resp = redirect('/interface/upload.html?_method=GET&message_type=error&message=文件格式错误')
            return resp
        else:
            length = ws.max_row
            header_row = ws[1]  # 假设表头在第一行
            header = [cell.value for cell in header_row]
            data = ws[2:length:1]
            if length>2:
                for i in data:
                    e_data = [cell.value for cell in i]
                    if m_sql.inser_id(file_name,header, e_data)==-1:
                        resp = redirect('/interface/upload.html?_method=GET&message_type=error&message=文件格式错误,请传输xlsx格式文件')
                        return resp
            else:
                e_data = [i.value for i in data]
                if m_sql.inser_id(file_name, header, e_data) == -1:
                    resp = redirect('/interface/upload.html?_method=GET&message_type=error&message=文件格式错误,注意时间格式与表头格式')
                    return resp

            file_path='file/'+f"{file_name}.xlsx"
            file_path = os.path.join(os.getcwd(), file_path)
            try:
                os.remove(file_path)
            except OSError as error:
                print(error)
            resp = redirect('/interface/experimenterCenter.html?_method=GET&message_type=success&message=上传成功')
            return resp

@app.route('/release', methods=['POST'])  # 处理实验传输
def handle_release():
        m_sql.build_sql()#数据库建立
        id=session.get('id')
        experimenter=request.form.get('experimenter')
        startDate= request.form.get('startDate')
        endDate= request.form.get('endDate')
        experiment = request.form.get('experiment')
        participantNumber = request.form.get('participantNumber')
        fee = request.form.get('fee')
        location = request.form.get('location')
        current_time = datetime.now().date()
        start_date_numeric = datetime.strptime(startDate, "%Y-%m-%d").strftime("%Y%m%d")
        end_date_numeric = datetime.strptime(endDate, "%Y-%m-%d").strftime("%Y%m%d")
        event='e'+id+experimenter+start_date_numeric+end_date_numeric+fee+location+experiment


        if current_time >= datetime.strptime(endDate, "%Y-%m-%d").date() :
            resp = redirect('/interface/recruit.html?_method=GET&message_type=error&message=截止日期小于当时日期')
            return resp
        if 'file' not in request.files:
            resp = redirect('/interface/experimenterCenter.html?_method=GET&message_type=error&message=文件传输错误1')
            return resp
        file_data=request.files['file']#接受实验数据
        if file_data.filename == '':
            resp = redirect('/interface/experimenterCenter.html?_method=GET&message_type=error&message=文件传输错误2')
            return resp

        UPLOAD_FOLDER = 'img'
        app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
        file_extension = os.path.splitext(file_data.filename)[1]  # 获取文件扩展名部分
        file_data.save(os.path.join(app.config['UPLOAD_FOLDER'], f"{event}{file_extension}"))#将文件保存在img文件下

        url=f"http://123.57.28.236:{port}/img/{event}{file_extension}"
        e_data = (id, experimenter, startDate, endDate, experiment, participantNumber, fee, location,url)
        if m_sql.inser_experiment(e_data)==-1:# 将信息保存在数据库中
            resp = redirect('/interface/experimenterCenter.html?_method=GET&message_type=error&message=时间格式有误')
            return resp
        #########################################################################################################
        #设置事件，到时间自动删除数据
        m_sql.create_event_experiment(id,experimenter,startDate,endDate,participantNumber,fee,location,experiment)
        file_path = 'img/' + f"{experiment}{file_extension}"
        file_path = os.path.join(os.getcwd(), file_path)
        #删除文件
        thread=threading.Thread(target=d_file_at_time,args=(file_path,endDate))
        thread.start()
        ##########################################################################################################
        resp = redirect('/interface/experimenterCenter.html?_method=GET&message_type=success&message=传输成功,请在右上角招募处查看上传实验信息')
        return resp


@app.route('/query_releasedInformation',methods=['POST'])  # 处理实验招募界面
def handle_query_release():
    # 接受返回的json
    packet = request.get_json()
    # 解析json
    page = packet['page']
    query_data = packet["filter"]
    # 获取待查询数据
    j_data = m_sql.select_sql_experiment()##占行
    if query_data == {}:
        j_data = m_sql.select_sql_experiment()# 得到数据
    else:
        k_l = list(query_data.keys())
        for i in k_l:
            if i == "date":
                j_data = m_sql.select_sql_experiment_time(query_data['date']['startDate'],
                                                       query_data['date']['endDate'])
    # 数据分块
    length = len(j_data) // 10
    remainder = len(j_data) % 10
    if length != 0:
        if remainder != 0:
            data_l = [j_data[i:i + 10] for i in range(0, length * 10, 10)]
            remainder_d = j_data[length * 10:]
            if page != length + 1:
                message = {"length": length + 1, "data": data_l[page - 1]}
            else:
                message = {"length": length + 1, "data": remainder_d}
        else:
            data_l = [j_data[i:i + 10] for i in range(0, length * 10, 10)]
            message = {"length": length, "data": data_l[page - 1]}
    else:
        if remainder != 0:
            message = {"length": 1, "data": j_data}
        else:
            message = {"length": 0, "data": []}


    response = make_response(json.dumps(message))
    response.headers['Server'] = 'test/1.0'
    response.headers['Content-Type'] = 'application/json; charset=utf-8'
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Cache-Control'] = 'no-store'
    response.headers['Access-Control-Allow-Headers'] = '*'
    response.headers['Access-Control-Allow-Methods'] = '*'
    return response



@app.route('/delete_releasedInformation',methods=['POST'])#处理停止实验
def handle_quit():
    id = session.get('id')
    j_data = request.get_json()
    experiment = j_data['experiment']
    experimenter = j_data['experimenter']
    start_date = j_data['startDate']
    end_date = j_data['endDate']
    number = j_data['participantNumber']
    fee = j_data['fee']
    location = j_data['location']
    url = j_data['URL']
    if m_sql.compare_sql_experiment(id,experimenter,start_date,end_date,number,fee,location,experiment)==1:
        m_sql.delete_sql_experiment(id,experimenter, start_date, end_date, number, fee, location,experiment)#删数据库里的数据
        #删文件
        filename=os.path.basename(url)
        file_path = 'img/' + filename
        file_path = os.path.join(os.getcwd(), file_path)
        os.remove(file_path)

        message={'message': "删除成功", 'messageType': "success"}
    else:
        message = {'message': "这不是你的实验", 'messageType': "error"}
    return jsonify(message)

if __name__ == '__main__':
    thread_h=threading.Thread(target=heart_check)
    thread_h.start()
    app.run(host='0.0.0.0', port=int(port))
